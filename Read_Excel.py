# -*- coding: utf-8 -*-
"""
Created on Tue Jun  2 15:55:14 2020

@author: M49835
Michael Yost 

This python script is used to generate JSON files from excel files. 
It may seem hard to follow at first but is just 3 dictionaries nested within each other. 
Nested level 1 = the excel sheetname (typically procedure name)
Nested level 2 = the excel column title (typcially the variable name ie. fault, subfault, etc)
Nested level 3 = N rows under the column title. (the specific variable names)
3 nested "for" loops for instance are:
    for loop ---nested level 1
        for loop -nested level 2
            for loop -nested level 3
This script will create level 3 nested JSON file from an excel file.  
"""

##Use Excel file to generate a JSON file. JSON file will be used for the python script

import xlrd
import openpyxl
from openpyxl import load_workbook
import json

workbook = load_workbook(filename = "C:/Landsat9_Procedures/trunk/DBVerif/Variable_Definitions.xlsx")

full_dict={}#create the nested level 1 dictionary. The completed dictionary/json file

for sheet in workbook:# nested level 1, the sheetname
    nested2_dict={}#create the nested level 2 dictionary
    workbook.active = sheet
    
    max_col = sheet.max_column
    max_row = sheet.max_row
    #print (sheet.title)
    #print (max_col)
    for c in range (1, max_col+1):#nested level 2, the column title dict
        
        row_array=[]
        title = sheet.cell(row=1, column=c).value#we get the column title
        
        nested3_dict ={}#nested level 3 the variables dict
        for r in range (2, max_row+1):
            
            row = sheet.cell(row=r, column=c).value
            
            if sheet.cell(row=r, column=c).value:
                row_array.append(row)#create our array of variable names
                
            nested3_dict.update({sheet.cell(row=1,column=c).value:row_array})
            
            nested2_dict.update(nested3_dict)
    
        
    full_dict.update ({sheet.title:nested2_dict})    
    
#print (full_dict)
full_dict = json.dumps(full_dict)
loaded_string = json.loads(full_dict)
print (loaded_string['L9_EPS_006']['neb'][0])

  
        
   